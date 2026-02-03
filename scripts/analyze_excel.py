"""
Script to get complete Excel sheet list - outputs to file
"""
from openpyxl import load_workbook
import json

wb = load_workbook('Form20_Localities_Pct (1).xlsx', read_only=True)
sheets = wb.sheetnames

output = []
output.append("="*60)
output.append("COMPLETE SHEET LIST")
output.append("="*60)
output.append(f"Total sheets: {len(sheets)}\n")

ac_sheets = []
other_sheets = []

for sheet in sheets:
    if sheet.startswith('AC_') and '_FINAL' in sheet:
        try:
            num = int(sheet.replace('AC_', '').replace('_FINAL', ''))
            ac_sheets.append((num, sheet))
        except:
            other_sheets.append(sheet)
    else:
        other_sheets.append(sheet)

# Sort AC sheets by number
ac_sheets.sort(key=lambda x: x[0])

output.append("Assembly Sheets Found:")
for num, name in ac_sheets:
    output.append(f"  AC {num:2d}: {name}")

output.append(f"\nTotal Assembly Sheets: {len(ac_sheets)}")

# Check for missing assemblies
expected = set(range(1, 31))
found = set(num for num, _ in ac_sheets)
missing = sorted(expected - found)

output.append(f"\nMissing Assemblies (1-30): {missing if missing else 'None'}")

output.append("\nOther Sheets:")
for sheet in other_sheets:
    output.append(f"  - {sheet}")

# Get column headers for each assembly sheet
output.append("\n" + "="*60)
output.append("COLUMN HEADERS BY ASSEMBLY")
output.append("="*60)

for num, sheet_name in ac_sheets[:5]:  # First 5 for sample
    ws = wb[sheet_name]
    headers = [cell.value for cell in list(ws.rows)[0] if cell.value]
    output.append(f"\n{sheet_name}:")
    output.append(f"  Columns ({len(headers)}): {headers}")

wb.close()

# Write to file
with open('excel_analysis.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(output))

print('\n'.join(output))
print("\nSaved to: excel_analysis.txt")
