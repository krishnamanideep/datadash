"""
Analyze structure of the three missing assembly Excel files
"""
from openpyxl import load_workbook
import os

files = ['AC_14_FINAL.xlsx', 'AC_16_FINAL.xlsx', 'AC_30_FINAL.xlsx']
assembly_names = {
    '14': 'Raj Bhavan',
    '16': 'Orleampeth', 
    '30': 'Yanam'
}

for filename in files:
    if not os.path.exists(filename):
        print(f"File not found: {filename}")
        continue
        
    print("="*60)
    print(f"FILE: {filename}")
    print("="*60)
    
    wb = load_workbook(filename, read_only=True)
    print(f"Sheets: {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Get headers (first row)
        rows = list(ws.rows)
        if rows:
            headers = [str(cell.value) if cell.value else '' for cell in rows[0]]
            print(f"\nSheet '{sheet_name}' - Columns ({len(headers)}):")
            print(f"  {headers}")
            
            # Count data rows
            data_rows = len(rows) - 1  # Exclude header
            print(f"  Data rows: {data_rows}")
            
            # Show sample of first data row
            if len(rows) > 1:
                first_row = [str(cell.value)[:30] if cell.value else 'NULL' for cell in rows[1]]
                print(f"  Sample row 1: {first_row[:10]}...")
    
    wb.close()
    print()
