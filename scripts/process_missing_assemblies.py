"""
Process the three missing assembly Excel files:
1. Merge 2021 and 11_16 sheets
2. Calculate vote percentages
3. Calculate party scores
4. Categorize booths (A/B/C)
5. Add coordinates based on known locality mappings
6. Export to JSON format matching existing data structure
"""
import json
from openpyxl import load_workbook
import re

# Region coordinates for geocoding
REGION_COORDS = {
    # AC_14 - Raj Bhavan (Puducherry urban)
    'RAJ BHAVAN': {'lat': 11.9316, 'lng': 79.8300},
    'PUDUCHERRY': {'lat': 11.9340, 'lng': 79.8330},
    'WHITE TOWN': {'lat': 11.9350, 'lng': 79.8310},
    'GOUBERT AVENUE': {'lat': 11.9300, 'lng': 79.8340},
    
    # AC_16 - Orleampeth (Puducherry)  
    'ORLEAMPETH': {'lat': 11.9050, 'lng': 79.8150},
    'ORLEANPET': {'lat': 11.9050, 'lng': 79.8150},
    
    # AC_30 - Yanam (Andhra Pradesh enclave)
    'YANAM': {'lat': 16.7333, 'lng': 82.2167},
    'FARAMPETA': {'lat': 16.7380, 'lng': 82.2150},
    'KONAPAPAPETA': {'lat': 16.7300, 'lng': 82.2200},
    'AGRAHARAM': {'lat': 16.7350, 'lng': 82.2180},
}

# Assembly-specific default coordinates
ASSEMBLY_DEFAULTS = {
    '14': {'lat': 11.9316, 'lng': 79.8300, 'name': 'Raj Bhavan'},
    '16': {'lat': 11.9050, 'lng': 79.8150, 'name': 'Orleampeth'},
    '30': {'lat': 16.7333, 'lng': 82.2167, 'name': 'Yanam'},
}

def extract_locality(ps_name):
    """Extract locality from polling station name"""
    if not ps_name:
        return 'Unknown'
    
    # Clean up the name
    ps_name = str(ps_name).upper()
    
    # Common patterns to remove
    remove_patterns = [
        r'GOVT\.?\s*',
        r'GOVERNMENT\s*',
        r'PRIMARY\s*SCHOOL\s*',
        r'HIGH\s*SCHOOL\s*',
        r'HIGHER\s*SECONDARY\s*SCHOOL\s*',
        r'MIDDLE\s*SCHOOL\s*',
        r'COMMUNITY\s*HALL\s*',
        r'\(NORTH\)',
        r'\(SOUTH\)',
        r'\(EAST\)',
        r'\(WEST\)',
        r'PUDUCHERRY-?\d*',
        r'YANAM-?\d*',
        r'-\d+$',
    ]
    
    for pattern in remove_patterns:
        ps_name = re.sub(pattern, '', ps_name)
    
    # Extract the locality (usually after 'STREET' or at end)
    words = ps_name.split()
    if len(words) >= 2:
        return ' '.join(words[-2:]).strip()
    return ps_name.strip() if ps_name.strip() else 'Unknown'

def get_coordinates(locality, assembly_num):
    """Get coordinates for a locality"""
    locality_upper = locality.upper()
    
    # Check for known localities
    for key, coords in REGION_COORDS.items():
        if key in locality_upper or locality_upper in key:
            # Add small random offset for different booths in same locality
            return coords['lat'], coords['lng']
    
    # Use assembly default with slight offset based on booth
    default = ASSEMBLY_DEFAULTS.get(assembly_num, ASSEMBLY_DEFAULTS['14'])
    return default['lat'], default['lng']

def safe_float(val, default=0):
    """Safely convert value to float"""
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default

def safe_int(val, default=0):
    """Safely convert value to int"""
    if val is None:
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default

def calculate_percentage(votes, total):
    """Calculate vote percentage"""
    if total and total > 0:
        return votes / total
    return 0

def process_assembly_14():
    """Process AC_14 - Raj Bhavan"""
    wb = load_workbook('AC_14_FINAL.xlsx')
    
    # Process 2021 sheet
    ws_2021 = wb['2021']
    data_2021 = {}
    for row in list(ws_2021.rows)[1:]:  # Skip header
        if row[0].value:
            ps_name = str(row[0].value)
            ps_num = safe_int(row[1].value)
            data_2021[ps_num] = {
                'PS_NO_2021': ps_name,
                'DMK_2021': safe_int(row[2].value),
                'AINRC_2021': safe_int(row[3].value),
                'MNM_2021': safe_int(row[4].value),
                'OTHERS_2021': safe_int(row[5].value),
                'NOTA_2021': safe_int(row[6].value),
                'POLLED_2021': safe_int(row[7].value),
                'VOTERS_2021': safe_int(row[8].value),
                'POLLED_%': safe_float(row[9].value),
            }
    
    # Process 11_16 sheet  
    ws_1116 = wb['11_16']
    data_1116 = {}
    for row in list(ws_1116.rows)[1:]:  # Skip header
        ps_num = safe_int(row[2].value)  # PS_NO_2016 is at index 2
        if ps_num:
            data_1116[ps_num] = {
                'AIADMK_2011': safe_int(row[3].value),
                'INC_2011': safe_int(row[4].value),
                'IND_2011': safe_int(row[5].value),  # ANNIBAL_NEHRU
                'OTHERS_2011': safe_int(row[6].value),
                'POLLED_2011': safe_int(row[7].value),
                'AINRC_2016': safe_int(row[9].value),
                'AIADMK_2016': safe_int(row[10].value),
                'INC_2016': safe_int(row[11].value),
                'OTHERS_2016': safe_int(row[12].value),
                'NOTA_2016': safe_int(row[13].value),
                'POLLED_2016': safe_int(row[9].value) + safe_int(row[10].value) + safe_int(row[11].value) + safe_int(row[12].value) + safe_int(row[13].value),
            }
    
    wb.close()
    
    # Merge and calculate
    result = []
    for ps_num, d2021 in data_2021.items():
        entry = {}
        entry['PS_NO_2021'] = d2021['PS_NO_2021']
        entry['LOCALITY_EXTRACTED'] = extract_locality(d2021['PS_NO_2021'])
        
        # Get coordinates with slight variation per booth
        base_lat, base_lng = get_coordinates(entry['LOCALITY_EXTRACTED'], '14')
        entry['Latitude'] = base_lat + (ps_num * 0.0003)  # Small offset per booth
        entry['Longitude'] = base_lng + (ps_num * 0.0002)
        
        # 2011 data
        d1116 = data_1116.get(ps_num, {})
        polled_2011 = d1116.get('POLLED_2011', 0)
        entry['AIADMK_2011'] = d1116.get('AIADMK_2011', 0)
        entry['AIADMK_2011_pct'] = calculate_percentage(entry['AIADMK_2011'], polled_2011)
        entry['INC_2011'] = d1116.get('INC_2011', 0)
        entry['INC_2011_pct'] = calculate_percentage(entry['INC_2011'], polled_2011)
        entry['IND_2011'] = d1116.get('IND_2011', 0)
        entry['IND_2011_pct'] = calculate_percentage(entry['IND_2011'], polled_2011)
        entry['OTHERS_2011'] = d1116.get('OTHERS_2011', 0)
        entry['OTHERS_2011_pct'] = calculate_percentage(entry['OTHERS_2011'], polled_2011)
        entry['POLLED_2011'] = polled_2011
        
        # 2016 data
        polled_2016 = d1116.get('POLLED_2016', 0)
        entry['AINRC_2016'] = d1116.get('AINRC_2016', 0)
        entry['AINRC_2016_pct'] = calculate_percentage(entry['AINRC_2016'], polled_2016)
        entry['AIADMK_2016'] = d1116.get('AIADMK_2016', 0)
        entry['AIADMK_2016_pct'] = calculate_percentage(entry['AIADMK_2016'], polled_2016)
        entry['INC_2016'] = d1116.get('INC_2016', 0)
        entry['INC_2016_pct'] = calculate_percentage(entry['INC_2016'], polled_2016)
        entry['OTHERS_2016'] = d1116.get('OTHERS_2016', 0)
        entry['OTHERS_2016_pct'] = calculate_percentage(entry['OTHERS_2016'], polled_2016)
        entry['NOTA_2016'] = d1116.get('NOTA_2016', 0)
        entry['NOTA_2016_pct'] = calculate_percentage(entry['NOTA_2016'], polled_2016)
        entry['POLLED_2016'] = polled_2016
        
        # 2021 data
        polled_2021 = d2021['POLLED_2021']
        entry['DMK_2021'] = d2021['DMK_2021']
        entry['DMK_2021_pct'] = calculate_percentage(entry['DMK_2021'], polled_2021)
        entry['AINRC_2021'] = d2021['AINRC_2021']
        entry['AINRC_2021_pct'] = calculate_percentage(entry['AINRC_2021'], polled_2021)
        entry['MNM_2021'] = d2021['MNM_2021']
        entry['MNM_2021_pct'] = calculate_percentage(entry['MNM_2021'], polled_2021)
        entry['OTHERS_2021'] = d2021['OTHERS_2021']
        entry['OTHERS_2021_pct'] = calculate_percentage(entry['OTHERS_2021'], polled_2021)
        entry['NOTA_2021'] = d2021['NOTA_2021']
        entry['NOTA_2021_pct'] = calculate_percentage(entry['NOTA_2021'], polled_2021)
        entry['POLLED_2021'] = polled_2021
        entry['VOTERS_2021'] = d2021['VOTERS_2021']
        entry['VOTERS_2021_pct'] = d2021['VOTERS_2021'] / polled_2021 if polled_2021 else 0
        entry['POLLED_%'] = d2021['POLLED_%']
        
        # Calculate scores (weighted average across years)
        scores = {}
        
        # Map parties to score calculation
        # 2011 weight: 0.2, 2016 weight: 0.3, 2021 weight: 0.5
        scores['AIADMK_SCORE'] = entry['AIADMK_2011_pct'] * 0.2 + entry['AIADMK_2016_pct'] * 0.3
        scores['INC_SCORE'] = entry['INC_2011_pct'] * 0.2 + entry['INC_2016_pct'] * 0.3
        scores['AINRC_SCORE'] = entry['AINRC_2016_pct'] * 0.3 + entry['AINRC_2021_pct'] * 0.5
        scores['DMK_SCORE'] = entry['DMK_2021_pct'] * 0.5
        scores['MNM_SCORE'] = entry['MNM_2021_pct'] * 0.5
        scores['OTHERS_SCORE'] = (entry['OTHERS_2011_pct'] * 0.2 + entry['OTHERS_2016_pct'] * 0.3 + entry['OTHERS_2021_pct'] * 0.5) / 3
        scores['IND_SCORE'] = entry['IND_2011_pct'] * 0.2
        
        # Normalize scores to sum to 1
        total_score = sum(scores.values())
        if total_score > 0:
            for key in scores:
                scores[key] = scores[key] / total_score
        
        entry.update(scores)
        
        # Find top score
        top_party = max(scores, key=scores.get)
        top_value = scores[top_party]
        party_name = top_party.replace('_SCORE', '')
        entry['TOP_SCORE_PARTY'] = f"{party_name} ({top_value*100:.2f}%)"
        
        # Categorize
        if top_value > 0.50:
            entry['TOP_SCORE_CATEGORY'] = 'A'
        elif top_value > 0.30:
            entry['TOP_SCORE_CATEGORY'] = 'B'
        else:
            entry['TOP_SCORE_CATEGORY'] = 'C'
        
        result.append(entry)
    
    return result

def process_assembly_16():
    """Process AC_16 - Orleampeth"""
    wb = load_workbook('AC_16_FINAL.xlsx')
    
    # Process 2021 sheet
    ws_2021 = wb['2021']
    data_2021 = {}
    for row in list(ws_2021.rows)[1:]:
        ps_num = safe_int(row[0].value)
        if ps_num:
            data_2021[ps_num] = {
                'PS_NO_2021': str(row[1].value) if row[1].value else '',
                'AIADMK_2021': safe_int(row[2].value),
                'DMK_2021': safe_int(row[3].value),
                'IND_2021': safe_int(row[4].value),  # KUPASWAMY
                'OTHERS_2021': safe_int(row[5].value),
                'NOTA_2021': safe_int(row[6].value),
                'POLLED_2021': safe_int(row[7].value),
                'VOTERS_2021': safe_int(row[8].value),
                'POLLED_%': safe_float(row[9].value),
            }
    
    # Process 11_16 sheet
    ws_1116 = wb['11_16']
    data_1116 = {}
    for row in list(ws_1116.rows)[1:]:
        ps_num = safe_int(row[1].value)  # PS_NO_2016 at index 1
        if ps_num:
            data_1116[ps_num] = {
                'DMK_2011': safe_int(row[4].value),
                'AINRC_2011': safe_int(row[5].value),
                'OTHERS_2011': safe_int(row[6].value),
                'POLLED_2011': safe_int(row[7].value),
                'DMK_2016': safe_int(row[9].value),
                'AINRC_2016': safe_int(row[10].value),
                'OTHERS_2016': safe_int(row[11].value),
                'NOTA_2016': safe_int(row[12].value),
                'POLLED_2016': safe_int(row[13].value),
            }
    
    wb.close()
    
    # Merge and calculate
    result = []
    for ps_num, d2021 in data_2021.items():
        entry = {}
        entry['PS_NO_2021'] = d2021['PS_NO_2021']
        entry['LOCALITY_EXTRACTED'] = extract_locality(d2021['PS_NO_2021'])
        
        base_lat, base_lng = get_coordinates(entry['LOCALITY_EXTRACTED'], '16')
        entry['Latitude'] = base_lat + (ps_num * 0.0003)
        entry['Longitude'] = base_lng + (ps_num * 0.0002)
        
        d1116 = data_1116.get(ps_num, {})
        polled_2011 = d1116.get('POLLED_2011', 0)
        
        entry['DMK_2011'] = d1116.get('DMK_2011', 0)
        entry['DMK_2011_pct'] = calculate_percentage(entry['DMK_2011'], polled_2011)
        entry['AINRC_2011'] = d1116.get('AINRC_2011', 0)
        entry['AINRC_2011_pct'] = calculate_percentage(entry['AINRC_2011'], polled_2011)
        entry['OTHERS_2011'] = d1116.get('OTHERS_2011', 0)
        entry['OTHERS_2011_pct'] = calculate_percentage(entry['OTHERS_2011'], polled_2011)
        entry['POLLED_2011'] = polled_2011
        
        polled_2016 = d1116.get('POLLED_2016', 0)
        entry['DMK_2016'] = d1116.get('DMK_2016', 0)
        entry['DMK_2016_pct'] = calculate_percentage(entry['DMK_2016'], polled_2016)
        entry['AINRC_2016'] = d1116.get('AINRC_2016', 0)
        entry['AINRC_2016_pct'] = calculate_percentage(entry['AINRC_2016'], polled_2016)
        entry['OTHERS_2016'] = d1116.get('OTHERS_2016', 0)
        entry['OTHERS_2016_pct'] = calculate_percentage(entry['OTHERS_2016'], polled_2016)
        entry['NOTA_2016'] = d1116.get('NOTA_2016', 0)
        entry['NOTA_2016_pct'] = calculate_percentage(entry['NOTA_2016'], polled_2016)
        entry['POLLED_2016'] = polled_2016
        
        polled_2021 = d2021['POLLED_2021']
        entry['AIADMK_2021'] = d2021['AIADMK_2021']
        entry['AIADMK_2021_pct'] = calculate_percentage(entry['AIADMK_2021'], polled_2021)
        entry['DMK_2021'] = d2021['DMK_2021']
        entry['DMK_2021_pct'] = calculate_percentage(entry['DMK_2021'], polled_2021)
        entry['IND_2021'] = d2021['IND_2021']
        entry['IND_2021_pct'] = calculate_percentage(entry['IND_2021'], polled_2021)
        entry['OTHERS_2021'] = d2021['OTHERS_2021']
        entry['OTHERS_2021_pct'] = calculate_percentage(entry['OTHERS_2021'], polled_2021)
        entry['NOTA_2021'] = d2021['NOTA_2021']
        entry['NOTA_2021_pct'] = calculate_percentage(entry['NOTA_2021'], polled_2021)
        entry['POLLED_2021'] = polled_2021
        entry['VOTERS_2021'] = d2021['VOTERS_2021']
        entry['VOTERS_2021_pct'] = d2021['VOTERS_2021'] / polled_2021 if polled_2021 else 0
        entry['POLLED_%'] = d2021['POLLED_%']
        
        # Calculate scores
        scores = {}
        scores['DMK_SCORE'] = entry['DMK_2011_pct'] * 0.2 + entry['DMK_2016_pct'] * 0.3 + entry['DMK_2021_pct'] * 0.5
        scores['AINRC_SCORE'] = entry['AINRC_2011_pct'] * 0.2 + entry['AINRC_2016_pct'] * 0.3
        scores['AIADMK_SCORE'] = entry['AIADMK_2021_pct'] * 0.5
        scores['IND_SCORE'] = entry['IND_2021_pct'] * 0.5
        scores['OTHERS_SCORE'] = (entry['OTHERS_2011_pct'] * 0.2 + entry['OTHERS_2016_pct'] * 0.3 + entry['OTHERS_2021_pct'] * 0.5)
        
        total_score = sum(scores.values())
        if total_score > 0:
            for key in scores:
                scores[key] = scores[key] / total_score
        
        entry.update(scores)
        
        top_party = max(scores, key=scores.get)
        top_value = scores[top_party]
        party_name = top_party.replace('_SCORE', '')
        entry['TOP_SCORE_PARTY'] = f"{party_name} ({top_value*100:.2f}%)"
        
        if top_value > 0.50:
            entry['TOP_SCORE_CATEGORY'] = 'A'
        elif top_value > 0.30:
            entry['TOP_SCORE_CATEGORY'] = 'B'
        else:
            entry['TOP_SCORE_CATEGORY'] = 'C'
        
        result.append(entry)
    
    return result

def process_assembly_30():
    """Process AC_30 - Yanam"""
    wb = load_workbook('AC_30_FINAL.xlsx')
    
    # Process 2021 sheet
    ws_2021 = wb['2021']
    data_2021 = {}
    for row in list(ws_2021.rows)[1:]:
        ps_name = str(row[0].value) if row[0].value else ''
        ps_num = safe_int(row[1].value)
        if ps_num or ps_name:
            key = ps_num if ps_num else len(data_2021) + 1
            data_2021[key] = {
                'PS_NO_2021': ps_name,
                'AINRC_2021': safe_int(row[2].value),
                'OTHERS_2021': safe_int(row[3].value),
                'NOTA_2021': safe_int(row[4].value),
                'POLLED_2021': safe_int(row[5].value),
                'VOTERS_2021': safe_int(row[6].value),
                'POLLED_%': safe_float(row[7].value),
            }
    
    # Process 11_16 sheet
    ws_1116 = wb['11_16']
    data_1116 = {}
    for row in list(ws_1116.rows)[1:]:
        ps_num = safe_int(row[2].value)  # PS_NO_2016 at index 2
        if ps_num:
            data_1116[ps_num] = {
                'AIADMK_2011': safe_int(row[4].value),
                'INC_2011': safe_int(row[5].value),
                'OTHERS_2011': safe_int(row[6].value),
                'POLLED_2011': safe_int(row[7].value),
                'AINRC_2016': safe_int(row[9].value),
                'INC_2016': safe_int(row[10].value),
                'OTHERS_2016': safe_int(row[11].value),
                'NOTA_2016': safe_int(row[12].value),
                'POLLED_2016': safe_int(row[13].value),
            }
    
    wb.close()
    
    # Merge and calculate
    result = []
    for ps_num, d2021 in data_2021.items():
        entry = {}
        entry['PS_NO_2021'] = d2021['PS_NO_2021']
        entry['LOCALITY_EXTRACTED'] = extract_locality(d2021['PS_NO_2021'])
        
        base_lat, base_lng = get_coordinates(entry['LOCALITY_EXTRACTED'], '30')
        entry['Latitude'] = base_lat + (ps_num * 0.0003)
        entry['Longitude'] = base_lng + (ps_num * 0.0002)
        
        d1116 = data_1116.get(ps_num, {})
        polled_2011 = d1116.get('POLLED_2011', 0)
        
        entry['AIADMK_2011'] = d1116.get('AIADMK_2011', 0)
        entry['AIADMK_2011_pct'] = calculate_percentage(entry['AIADMK_2011'], polled_2011)
        entry['INC_2011'] = d1116.get('INC_2011', 0)
        entry['INC_2011_pct'] = calculate_percentage(entry['INC_2011'], polled_2011)
        entry['OTHERS_2011'] = d1116.get('OTHERS_2011', 0)
        entry['OTHERS_2011_pct'] = calculate_percentage(entry['OTHERS_2011'], polled_2011)
        entry['POLLED_2011'] = polled_2011
        
        polled_2016 = d1116.get('POLLED_2016', 0)
        entry['AINRC_2016'] = d1116.get('AINRC_2016', 0)
        entry['AINRC_2016_pct'] = calculate_percentage(entry['AINRC_2016'], polled_2016)
        entry['INC_2016'] = d1116.get('INC_2016', 0)
        entry['INC_2016_pct'] = calculate_percentage(entry['INC_2016'], polled_2016)
        entry['OTHERS_2016'] = d1116.get('OTHERS_2016', 0)
        entry['OTHERS_2016_pct'] = calculate_percentage(entry['OTHERS_2016'], polled_2016)
        entry['NOTA_2016'] = d1116.get('NOTA_2016', 0)
        entry['NOTA_2016_pct'] = calculate_percentage(entry['NOTA_2016'], polled_2016)
        entry['POLLED_2016'] = polled_2016
        
        polled_2021 = d2021['POLLED_2021']
        entry['AINRC_2021'] = d2021['AINRC_2021']
        entry['AINRC_2021_pct'] = calculate_percentage(entry['AINRC_2021'], polled_2021)
        entry['OTHERS_2021'] = d2021['OTHERS_2021']
        entry['OTHERS_2021_pct'] = calculate_percentage(entry['OTHERS_2021'], polled_2021)
        entry['NOTA_2021'] = d2021['NOTA_2021']
        entry['NOTA_2021_pct'] = calculate_percentage(entry['NOTA_2021'], polled_2021)
        entry['POLLED_2021'] = polled_2021
        entry['VOTERS_2021'] = d2021['VOTERS_2021']
        entry['VOTERS_2021_pct'] = d2021['VOTERS_2021'] / polled_2021 if polled_2021 else 0
        entry['POLLED_%'] = d2021['POLLED_%']
        
        # Calculate scores
        scores = {}
        scores['AIADMK_SCORE'] = entry['AIADMK_2011_pct'] * 0.2
        scores['INC_SCORE'] = entry['INC_2011_pct'] * 0.2 + entry['INC_2016_pct'] * 0.3
        scores['AINRC_SCORE'] = entry['AINRC_2016_pct'] * 0.3 + entry['AINRC_2021_pct'] * 0.5
        scores['OTHERS_SCORE'] = (entry['OTHERS_2011_pct'] * 0.2 + entry['OTHERS_2016_pct'] * 0.3 + entry['OTHERS_2021_pct'] * 0.5)
        
        total_score = sum(scores.values())
        if total_score > 0:
            for key in scores:
                scores[key] = scores[key] / total_score
        
        entry.update(scores)
        
        top_party = max(scores, key=scores.get)
        top_value = scores[top_party]
        party_name = top_party.replace('_SCORE', '')
        entry['TOP_SCORE_PARTY'] = f"{party_name} ({top_value*100:.2f}%)"
        
        if top_value > 0.50:
            entry['TOP_SCORE_CATEGORY'] = 'A'
        elif top_value > 0.30:
            entry['TOP_SCORE_CATEGORY'] = 'B'
        else:
            entry['TOP_SCORE_CATEGORY'] = 'C'
        
        result.append(entry)
    
    return result

def main():
    print("Processing missing assembly data...")
    
    # Load existing JSON
    with open('Form20_Localities_Pct.json', 'r', encoding='utf-8') as f:
        existing_data = json.load(f)
    
    print(f"Loaded existing data with {len(existing_data)} keys")
    
    # Process AC_14
    print("\nProcessing AC_14 (Raj Bhavan)...")
    ac_14_data = process_assembly_14()
    print(f"  Processed {len(ac_14_data)} entries")
    existing_data['AC_14_FINAL'] = ac_14_data
    
    # Process AC_16
    print("\nProcessing AC_16 (Orleampeth)...")
    ac_16_data = process_assembly_16()
    print(f"  Processed {len(ac_16_data)} entries")
    existing_data['AC_16_FINAL'] = ac_16_data
    
    # Process AC_30
    print("\nProcessing AC_30 (Yanam)...")
    ac_30_data = process_assembly_30()
    print(f"  Processed {len(ac_30_data)} entries")
    existing_data['AC_30_FINAL'] = ac_30_data
    
    # Save updated JSON
    print("\nSaving updated JSON...")
    with open('Form20_Localities_Pct.json', 'w', encoding='utf-8') as f:
        json.dump(existing_data, f, indent=1)
    
    print("Done!")
    print(f"\nTotal keys in updated JSON: {len(existing_data)}")
    
    # Print summary
    print("\n" + "="*60)
    print("PROCESSING SUMMARY")
    print("="*60)
    
    for ac_num, data in [('14', ac_14_data), ('16', ac_16_data), ('30', ac_30_data)]:
        print(f"\nAC_{ac_num} ({ASSEMBLY_DEFAULTS[ac_num]['name']}):")
        print(f"  Total booths: {len(data)}")
        
        # Category breakdown
        cat_a = sum(1 for e in data if e['TOP_SCORE_CATEGORY'] == 'A')
        cat_b = sum(1 for e in data if e['TOP_SCORE_CATEGORY'] == 'B')
        cat_c = sum(1 for e in data if e['TOP_SCORE_CATEGORY'] == 'C')
        print(f"  Category A (>50%): {cat_a} booths")
        print(f"  Category B (30-50%): {cat_b} booths")
        print(f"  Category C (<30%): {cat_c} booths")
        
        # Top party breakdown
        parties = {}
        for e in data:
            party = e['TOP_SCORE_PARTY'].split()[0]
            parties[party] = parties.get(party, 0) + 1
        print(f"  Dominant parties: {parties}")

if __name__ == '__main__':
    main()
